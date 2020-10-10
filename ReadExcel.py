import openpyxl

path="C://Users/HP/Google Drive/Learning/Selenium/readData.xlsx"

workbook=openpyxl.load_workbook(path) #also use:workbook.get_sheet_by_name("Sheet1")

sheet=workbook.active

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="               ")
    print() #print in the next line
